"""XLSX extractor backed by ``openpyxl``.

``openpyxl`` is the light, purpose-built choice (``pandas`` would drag in
numpy for no benefit on a pure-parse path). Lazily imported; the error
message points at the ``[ingest-xlsx]`` extra. One section per worksheet;
rows are serialized as a Markdown table.
"""

from __future__ import annotations

import io
from typing import Any, ClassVar

from ...exceptions import KaguraIngestError
from .._types import ExtractedContent, ExtractedSection
from ._util import MAX_TOTAL_TEXT_CHARS as _MAX_TOTAL_TEXT_CHARS
from ._util import filename_title

# XLSX is a ZIP container — cap sheets, rows-per-sheet, total serialized text,
# and the materialized table size to defuse decompression bombs and pathological
# grids. The cell cap bounds the width-padding allocation (one wide row would
# otherwise force every row to be padded to that width).
_MAX_SHEETS = 1_000
_MAX_ROWS_PER_SHEET = 100_000
_MAX_TABLE_CELLS = 2_000_000  # width × row count after padding


def _cell_str(value: Any) -> str:
    """Stringify a cell for a Markdown table: escape pipes and flatten any
    newline variant (``\\r\\n``, ``\\r``, ``\\n``) to a space so a multi-line
    cell can't break the single-line table row."""
    if value is None:
        return ""
    return str(value).replace("|", r"\|").replace("\r\n", " ").replace("\r", " ").replace("\n", " ")


def _load_openpyxl() -> Any:
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as e:
        raise KaguraIngestError(
            "openpyxl is not installed. Install with: pip install 'kagura-memory[ingest-xlsx]'"
        ) from e
    return openpyxl


class XlsxExtractor:
    """Structural XLSX extractor.

    Each worksheet becomes one section whose heading is the sheet name and
    whose body is a Markdown table of the populated cells. ``page_range`` is
    ``None`` — spreadsheets are not paginated.
    """

    supports: ClassVar[frozenset[str]] = frozenset(
        {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    def extract(self, source: bytes, source_uri: str | None = None) -> ExtractedContent:
        openpyxl = _load_openpyxl()
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)
        except Exception as e:  # noqa: BLE001 - re-raised as our domain error
            raise KaguraIngestError(f"failed to open XLSX: {e}") from e

        try:
            sheets = workbook.worksheets
            if len(sheets) > _MAX_SHEETS:
                raise KaguraIngestError(
                    f"XLSX sheet count {len(sheets)} exceeds limit {_MAX_SHEETS}"
                )
            sections = self._sections(sheets)
        finally:
            workbook.close()

        title = self._title(source_uri)
        return ExtractedContent(title=title, sections=sections, images=[], page_count=None)

    @classmethod
    def _sections(cls, sheets: list[Any]) -> list[ExtractedSection]:
        sections: list[ExtractedSection] = []
        total_chars = 0
        for sheet in sheets:
            table, total_chars = cls._sheet_to_markdown(sheet, total_chars)
            if not table.strip():
                continue
            sections.append(
                ExtractedSection(
                    heading=str(sheet.title),
                    body_text=table,
                    page_range=None,
                    depth=1,
                    anchor=str(sheet.title),
                )
            )
        return sections

    @staticmethod
    def _sheet_to_markdown(sheet: Any, total_chars: int) -> tuple[str, int]:
        """Serialize a sheet to a Markdown table, accumulating ``total_chars``.

        The document-wide character budget is checked per row (not per sheet)
        so a runaway sheet is rejected before its full table is materialized.
        Returns the table text and the updated running total.
        """
        # Fast fail on a hostile DECLARED grid before iter_rows streams it: a
        # far-flung cell (or formatting) can inflate max_column/max_row so the
        # iterator allocates wide tuples row-by-row. This mirrors the post-loop
        # width × rows cap (clamped to the row limit we'd actually read), so it
        # rejects the same sheets — just earlier and in O(1). max_* may be None
        # when the dimension is undeclared; skip the pre-check then.
        max_col = sheet.max_column
        max_row = sheet.max_row
        if max_col and max_row and max_col * min(max_row, _MAX_ROWS_PER_SHEET) > _MAX_TABLE_CELLS:
            raise KaguraIngestError(
                f"XLSX sheet {sheet.title!r} declared grid {max_col}x{max_row} exceeds "
                f"{_MAX_TABLE_CELLS} cells (decompression bomb?)"
            )

        rows: list[list[str]] = []
        cell_count = 0
        # Bound iteration at the source: openpyxl stops after _MAX_ROWS_PER_SHEET
        # + 1 rows instead of streaming a hostile max_row. The "+1" still lets a
        # genuinely over-limit sheet produce the row that trips the raise below
        # (so oversized input fails with KaguraIngestError rather than silently
        # truncating).
        for row_idx, row in enumerate(
            sheet.iter_rows(values_only=True, max_row=_MAX_ROWS_PER_SHEET + 1)
        ):
            if row_idx >= _MAX_ROWS_PER_SHEET:
                raise KaguraIngestError(
                    f"XLSX sheet {sheet.title!r} exceeds {_MAX_ROWS_PER_SHEET} rows "
                    "(decompression bomb?)"
                )
            cells = [_cell_str(v) for v in row]
            # Drop trailing empty cells so a sparse sheet does not emit a wall
            # of empty columns.
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                # Cap the unpadded cell count DURING the loop so a wide-sparse
                # sheet (each row carrying a trailing cell at a far column)
                # cannot materialize a giant `rows` list before the post-loop
                # padding check runs.
                cell_count += len(cells)
                if cell_count > _MAX_TABLE_CELLS:
                    raise KaguraIngestError(
                        f"XLSX sheet {sheet.title!r} exceeds {_MAX_TABLE_CELLS} cells "
                        "(decompression bomb?)"
                    )
                total_chars += sum(len(c) for c in cells)
                if total_chars > _MAX_TOTAL_TEXT_CHARS:
                    raise KaguraIngestError(
                        f"XLSX total text exceeds {_MAX_TOTAL_TEXT_CHARS} chars "
                        "(decompression bomb?)"
                    )
                rows.append(cells)
        if not rows:
            return "", total_chars

        width = max(len(r) for r in rows)
        # Also bound the PADDED table before materializing it: a single wide row
        # forces every row to be padded to `width`, so width × rows can explode
        # even when the unpadded cell count and total_chars stay under their caps.
        if width * len(rows) > _MAX_TABLE_CELLS:
            raise KaguraIngestError(
                f"XLSX sheet {sheet.title!r} table size {width}x{len(rows)} exceeds "
                f"{_MAX_TABLE_CELLS} cells (decompression bomb?)"
            )
        rows = [r + [""] * (width - len(r)) for r in rows]
        header = rows[0]
        lines = [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(["---"] * width) + " |",
        ]
        for r in rows[1:]:
            lines.append("| " + " | ".join(r) + " |")
        return "\n".join(lines), total_chars

    @staticmethod
    def _title(source_uri: str | None) -> str | None:
        return filename_title(source_uri)
